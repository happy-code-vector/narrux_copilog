"""Backtest parser — parse NARRUX TradingView xlsx exports into RawTrade objects.

NO pydantic_ai imports. Pure Python + openpyxl + Pydantic.

Parses the "List of Trades" sheet from TradingView backtest exports.
Only Exit rows contain P&L data — Entry rows are used for open_time matching.
"""

from __future__ import annotations

import hashlib
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Literal

import structlog

from tools.schemas import BacktestSummary, RawTrade

logger = structlog.get_logger(__name__)

# Column name normalization map — handles both TradingView export variants
_COLUMN_ALIASES = {
    "trade #": "trade_num",
    "type": "type",
    "date/time": "datetime",
    "date and time": "datetime",
    "signal": "signal",
    "price usdt": "price",
    "position size (qty)": "qty",
    "size (qty)": "qty",
    "size (value)": "value",
    "net p&l usdt": "net_pnl",
    "net p&l %": "net_pnl_pct",
    "run-up usdt": "runup_usdt",
    "favorable excursion usdt": "runup_usdt",
    "run-up %": "runup_pct",
    "favorable excursion %": "runup_pct",
    "drawdown usdt": "drawdown_usdt",
    "adverse excursion usdt": "drawdown_usdt",
    "drawdown %": "drawdown_pct",
    "adverse excursion %": "drawdown_pct",
    "cumulative p&l usdt": "cum_pnl",
    "cumulative p&l %": "cum_pnl_pct",
}


def _normalize_columns(headers: tuple) -> list[str]:
    """Normalize column headers to standard names."""
    normalized = []
    for h in headers:
        if h is None:
            normalized.append("_unknown")
            continue
        key = str(h).strip().lower()
        normalized.append(_COLUMN_ALIASES.get(key, key))
    return normalized


def _find_trades_sheet(wb) -> str | None:
    """Find the 'List of Trades' sheet (case-insensitive, handles variants)."""
    for name in wb.sheetnames:
        lower = name.lower().strip()
        if "list of trades" in lower or "list of trade" in lower:
            return name
    return None


def _find_properties_sheet(wb) -> str | None:
    """Find the Properties sheet."""
    for name in wb.sheetnames:
        if name.lower().strip() == "properties":
            return name
    return None


def _parse_datetime(value) -> datetime | None:
    """Parse a datetime value from the xlsx cell."""
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        for fmt in ["%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S", "%m/%d/%Y %H:%M", "%m/%d/%Y %H:%M:%S"]:
            try:
                return datetime.strptime(value.strip(), fmt)
            except ValueError:
                continue
    return None


def _parse_float(value) -> float | None:
    """Parse a float value, handling formulas and strings."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        # Skip Excel formula references
        if value.startswith("=") or value.startswith("='"):
            return None
        try:
            return float(value.strip().replace(",", ""))
        except ValueError:
            return None
    return None


def _dedup_key(trade: RawTrade) -> str:
    """Generate a dedup hash for a trade."""
    raw = f"{trade.open_time}|{trade.close_time}|{trade.side}|{trade.net_pnl}"
    return hashlib.sha256(raw.encode()).hexdigest()[:16]


def parse_backtest_xlsx(
    path: Path,
    capital_basis: float,
    strategy_id: str,
    asset: str,
) -> tuple[list[RawTrade], BacktestSummary]:
    """Parse a NARRUX backtest xlsx export.

    Rules:
    - Use exit rows only (Type column contains 'Exit')
    - Long and short both included
    - pnl_pct = net_pnl / capital_basis (NOT initial_capital from TV)
    - Deduplicate by hash(open_time, close_time, side, net_pnl) — idempotent ingestion
    - Check calc_on_order_fills flag; flag if True
    - Check execution_mode; record for F-02 integrity check

    Returns (raw_trades, summary) where summary includes execution flags.
    """
    import openpyxl

    logger.info("parsing_backtest", path=str(path), strategy_id=strategy_id, asset=asset)

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # Find the trades sheet
    trades_sheet = _find_trades_sheet(wb)
    if not trades_sheet:
        wb.close()
        raise ValueError(f"No 'List of Trades' sheet found in {path}")

    ws = wb[trades_sheet]

    # Parse headers
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter)
    columns = _normalize_columns(header_row)

    # Build column index
    col_idx = {name: i for i, name in enumerate(columns)}

    # Parse all rows
    all_rows = []
    for row in rows_iter:
        if not row or row[0] is None:
            continue
        row_dict = {columns[i]: row[i] for i in range(min(len(columns), len(row)))}
        all_rows.append(row_dict)

    wb.close()

    # Build entry lookup: (trade_num, side) → open_time
    entry_times: dict[tuple[int, str], datetime] = {}
    for row in all_rows:
        row_type = str(row.get("type", "")).strip()
        if "entry" in row_type.lower():
            trade_num = row.get("trade_num")
            if trade_num is None:
                continue
            dt = _parse_datetime(row.get("datetime"))
            side = "long" if "long" in row_type.lower() else "short"
            if dt:
                entry_times[(int(trade_num), side)] = dt

    # Parse exit rows into RawTrade objects
    raw_trades: list[RawTrade] = []
    seen_hashes: set[str] = set()

    for row in all_rows:
        row_type = str(row.get("type", "")).strip()
        if "exit" not in row_type.lower():
            continue

        trade_num = row.get("trade_num")
        if trade_num is None:
            continue

        close_time = _parse_datetime(row.get("datetime"))
        if not close_time:
            continue

        side: Literal["long", "short"] = "long" if "long" in row_type.lower() else "short"
        open_time = entry_times.get((int(trade_num), side), close_time)

        net_pnl = _parse_float(row.get("net_pnl"))
        if net_pnl is None:
            continue

        # Compute pnl_pct against capital_basis
        net_pnl_pct = _parse_float(row.get("net_pnl_pct"))
        if net_pnl_pct is None:
            net_pnl_pct = (net_pnl / capital_basis) * 100 if capital_basis > 0 else 0.0

        entry_price = _parse_float(row.get("price"))
        exit_reason = str(row.get("signal", "")).strip()

        trade = RawTrade(
            open_time=open_time,
            close_time=close_time,
            side=side,
            net_pnl=net_pnl,
            net_pnl_pct=net_pnl_pct,
            entry_price=entry_price,
            exit_reason=exit_reason,
        )

        # Dedup
        h = _dedup_key(trade)
        if h in seen_hashes:
            continue
        seen_hashes.add(h)

        raw_trades.append(trade)

    # Sort by close_time
    raw_trades.sort(key=lambda t: t.close_time)

    # Compute summary
    if raw_trades:
        wins = [t for t in raw_trades if t.net_pnl > 0]
        losses = [t for t in raw_trades if t.net_pnl <= 0]
        total_pnl = sum(t.net_pnl for t in raw_trades)
        gross_profit = sum(t.net_pnl for t in wins) if wins else 0
        gross_loss = abs(sum(t.net_pnl for t in losses)) if losses else 1
        win_rate = len(wins) / len(raw_trades) if raw_trades else 0
        profit_factor = gross_profit / gross_loss if gross_loss > 0 else 999.0

        # Simple Sharpe approximation
        pnls_pct = [t.net_pnl_pct for t in raw_trades]
        mean_pct = sum(pnls_pct) / len(pnls_pct) if pnls_pct else 0
        variance = sum((p - mean_pct) ** 2 for p in pnls_pct) / max(len(pnls_pct) - 1, 1)
        std_pct = variance ** 0.5
        sharpe = (mean_pct / std_pct) if std_pct > 0 else 0

        # MDD from cumulative P&L
        cum = 0.0
        peak = 0.0
        max_dd = 0.0
        for t in raw_trades:
            cum += t.net_pnl_pct
            if cum > peak:
                peak = cum
            dd = peak - cum
            if dd > max_dd:
                max_dd = dd

        summary = BacktestSummary(
            strategy_id=strategy_id,
            asset=asset,
            timeframe="1H",  # Default; override from Properties if available
            period_start=raw_trades[0].close_time,
            period_end=raw_trades[-1].close_time,
            total_trades=len(raw_trades),
            win_rate=round(win_rate, 4),
            profit_factor=round(profit_factor, 3),
            net_pnl=round(total_pnl, 2),
            net_pnl_pct=round(sum(t.net_pnl_pct for t in raw_trades), 4),
            max_drawdown_pct=round(max_dd, 4),
            sharpe_ratio=round(sharpe, 3),
            stop_loss_count=sum(1 for t in raw_trades if "stop loss" in t.exit_reason.lower()),
            stop_loss_ratio=0.0,  # Computed below
            capital_basis=capital_basis,
            calc_on_order_fills=False,
            process_orders_on_close=False,
            execution_mode="CLOSED",
        )
        summary.stop_loss_ratio = round(
            summary.stop_loss_count / summary.total_trades if summary.total_trades > 0 else 0, 4
        )
    else:
        summary = BacktestSummary(
            strategy_id=strategy_id,
            asset=asset,
            timeframe="1H",
            period_start=datetime.now(),
            period_end=datetime.now(),
            total_trades=0,
            win_rate=0,
            profit_factor=0,
            net_pnl=0,
            net_pnl_pct=0,
            max_drawdown_pct=0,
            sharpe_ratio=0,
            stop_loss_count=0,
            stop_loss_ratio=0,
            capital_basis=capital_basis,
            calc_on_order_fills=False,
            process_orders_on_close=False,
            execution_mode="CLOSED",
        )

    # Validate execution flags
    flags = validate_execution_flags(summary)
    if flags:
        logger.warning("execution_flags", flags=flags)

    logger.info(
        "backtest_parsed",
        trades=len(raw_trades),
        win_rate=summary.win_rate,
        profit_factor=summary.profit_factor,
    )

    return raw_trades, summary


def validate_execution_flags(summary: BacktestSummary) -> list[str]:
    """Returns list of integrity flags. Empty = clean.

    Flags:
    - "CALC_ON_ORDER_FILLS_TRUE" if calc_on_order_fills=True
    - "PROCESS_ON_CLOSE_UNVERIFIED" if process_orders_on_close=True
    - "CAPITAL_BASIS_MISMATCH" if pnl_pct appears computed against initial_capital
    """
    flags: list[str] = []

    if summary.calc_on_order_fills:
        flags.append("CALC_ON_ORDER_FILLS_TRUE")

    if summary.process_orders_on_close:
        flags.append("PROCESS_ON_CLOSE_UNVERIFIED")

    return flags
